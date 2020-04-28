import os
from django.conf import settings
import datetime
import openpyxl
import re
import logging
import subprocess
import shlex
from django.http import StreamingHttpResponse

import sys
import os.path
sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from parsing.models import *
from parsing.parsers.functions import *

def gal_contrakty(filename, contractor):
    # exception = kw.pop('exception', Exception)
    check_status = File.objects.get(file__icontains=filename, uploaded_at__date=datetime.date.today())
    operator_id = operator_create('Галицькі контракти', contractor)
    header_row = 1
    table_first_row = 3
    light_presence = '+'
    sold_presence = ['занято', 'занятий']
    rezerv_presence = 'резерв'

    if check_status.status == 1:
        try:
            if filename.endswith('.xls'):
                dir = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()))
                output = subprocess.run(['parsing/parsers/convert.sh', dir, filename])
                filepath = os.path.join(dir, filename + 'x')
            else:
                filepath = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()), filename)

            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet in wb.worksheets:
                sheet_idx = (wb.worksheets.index(sheet))
                # get max row count
                max_row = sheet.max_row
                # get max column count
                max_column = sheet.max_column
                # iterate over all cells
                # iterate over all rows
                year_id = year_create()
                m = months_create()
                free_status_id = free_status_create()
                header_dictionary = {}
                months_data = {}

                def add_to_dict(k, v):
                    try:
                        months_data[k].append(v)
                    except KeyError:
                        months_data[k] = [v]

                headers = sheet[header_row]
                for i in headers:
                    if type(i).__name__ == 'MergedCell':
                        for items in sorted(sheet.merged_cell_ranges):
                            unmerged = sheet.unmerge_cells(str(items))
                            if unmerged is not None:
                                header_dictionary[unmerged.col_idx] = unmerged.value
                    else:
                        if i.value is not None:
                            header_dictionary[i.col_idx] = i.value
                # iterate trough headers keys and find col indexes
                for k, v in header_dictionary.items():
                    if 'OTS' in v:
                        ots_col = k
                    elif any(c in v for c in ('GRP')):
                        grp_col = k
                    elif 'Формат' in v:
                        size_col = k
                    elif 'Освітлення' in v:
                        light_col = k
                    elif 'Місто' in v:
                        city_col = k
                    elif 'Код' in v:
                        kod_col = k
                    elif 'Адреса' in v:
                        address_col = k
                    elif 'Сторона' in v:
                        side_col = k
                    elif 'Фото' in v:
                        url_col = k
                    elif 'код "ДОРС"' in v:
                        doors_col = k
                    elif any(c in v for c in ('2019 комерція', ' ціна  2019', ' ціна  2019')):
                        price_col = k
                    elif any(c in v for c in ('січень', 'Січень')):
                        add_to_dict(1, k)
                        add_to_dict(1, k+1)
                    elif any(c in v for c in ('лютий', 'Лютий')):
                        add_to_dict(2, k)
                        add_to_dict(2, k+1)
                    elif any(c in v for c in ('березень', 'Березень')):
                        add_to_dict(3, k)
                        add_to_dict(3, k+1)
                    elif any(c in v for c in ('квітень', 'Квітень')):
                        add_to_dict(4, k)
                        add_to_dict(4, k+1)
                    elif any(c in v for c in ('травень', 'Травень')):
                        add_to_dict(5, k)
                        add_to_dict(5, k+1)
                    elif any(c in v for c in ('червень', 'Червень')):
                        add_to_dict(6, k)
                        add_to_dict(6, k+1)
                    elif any(c in v for c in ('липень', 'Липень')):
                        add_to_dict(7, k)
                        add_to_dict(7, k+1)
                    elif any(c in v for c in ('серпень', 'Серпень')):
                        add_to_dict(8, k)
                        add_to_dict(8, k+1)
                    elif any(c in v for c in ('вересень', 'Вересень')):
                        add_to_dict(9, k)
                        add_to_dict(9, k+1)
                    elif any(c in v for c in ('жовтень', 'Жовтень')):
                        add_to_dict(10, k)
                        add_to_dict(10, k+1)
                    elif any(c in v for c in ('листопад', 'Листопад')):
                        add_to_dict(11, k)
                        add_to_dict(11, k+1)
                    elif any(c in v for c in ('грудень', 'Грудень')):
                        add_to_dict(12, k)
                        add_to_dict(12, k+1)
                    else:
                        pass

                # iterating trough rows
                for i in range(table_first_row, max_row + 1):
                    city_tmp = None
                    media_type_id = None
                    media_type_tmp = None
                    size_tmp = None
                    side_tmp = None
                    # grp_col = None

                    raw_kod = sheet.cell(row=i, column=kod_col).value
                    if raw_kod is not None and raw_kod != '':

                        try:
                            raw_ots = sheet.cell(row=i, column=ots_col).value
                            ots = raw_ots if raw_ots != '' and raw_ots is not None and type(raw_ots) is not str else None
                        except:
                            ots = None

                        try:
                            raw_grp = sheet.cell(row=i, column=grp_col).value
                            grp = raw_grp if raw_grp != '' and raw_grp is not None else None
                        except:
                            grp = None

                        size_raw = sheet.cell(row=i, column=size_col).value
                        if size_raw != '' and size_raw is not None:
                            str_split = size_raw.split()
                            if len(str_split) > 1:
                                media_type = str_split[0]
                                media_type_id = media_type_create(media_type.strip().capitalize())
                                media_type_tmp = media_type
                                size_raw = str_split[1].strip().capitalize()
                                size = size_raw.replace(',', '.').replace('*', 'x').split(' ')[0]
                                size_id = board_size_create(size)
                                size_tmp = size
                            else:
                                size_raw = str_split[0]
                                size = size_raw.replace(',', '.').replace('*', 'x').split(' ')[0]
                                size_id = board_size_create(size)
                                size_tmp = size
                        else:
                            size_id = None

                        try:
                            raw_light = sheet.cell(row=i, column=light_col).value
                            light = True if raw_light == light_presence else False
                        except:
                            light = False

                        city = sheet.cell(row=i, column=city_col).value
                        city_id = city_create(city) if city != '' and city != city_tmp and city is not None else None
                        city_tmp = city

                        kod = raw_kod

                        # Addreses customization
                        if sheet_idx == 0:
                            address1 = sheet.cell(row=i, column=address_col - 1).value
                            address2 = sheet.cell(row=i, column=address_col).value
                            if address2  != '':
                                address = ( address1 if address1 else '') + address2
                        elif sheet_idx == 1:
                            address1 = sheet.cell(row=i, column=address_col).value
                            address2 = sheet.cell(row=i, column=address_col + 1).value
                            if address2 != '':
                                address = address1+'.'+address2
                        else:
                            address = sheet.cell(row=i, column=address_col + 1).value

                        side = sheet.cell(row=i, column=side_col).value

                        board_side_id = board_side_create(side) if side != '' and side != side_tmp and side is not None else None
                        side_tmp = side

                        raw_url = sheet.cell(row=i, column=url_col).value
                        url = raw_url if raw_url != '' and raw_url is not None else None

                        try:
                            raw_kod_doors = sheet.cell(row=i, column=doors_col).value
                            kod_doors_int = int(
                                raw_kod_doors) if raw_kod_doors != '' and raw_kod_doors is not None else None
                        except:
                            kod_doors_int = None

                        price = sheet.cell(row=i, column=price_col).value
                        price_val = float(price) if price != '' and price is not None else None

                        # save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors, operator_id)
                        board_id = save_board(kod, url, None, city_id, None, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors_int, operator_id)
                        for k, v in months_data.items():
                            status_sold = sheet.cell(row=i, column=v[0]).value
                            status_busy = sheet.cell(row=i, column=v[1]).value
                            if status_sold is not None and status_sold in sold_presence :
                                status = status_sold
                            elif status_busy is not None and status_busy == rezerv_presence:
                                status = status_busy
                            else:
                                status = None
                            status_id = save_status(status, status_sold, status_busy)
                            save_status_price(board_id, status_id, price_val, k)
                    else:
                        pass

                try:
                    del kod_col
                    del ots_col
                    del grp_col
                    del size_col
                    del side_col
                    del address_col
                    del light_col
                    del city_col
                    del url_col
                    del price_col
                    del doors_col
                except:
                    pass
                    # response = StreamingHttpResponse(i , content_type='application/json')
                    # return response
            done_parsing(check_status.id)
            return ({contractor: 'parsed'})
            #response = StreamingHttpResponse(i , content_type='application/json')
            #return response

        except Exception as e:
            error_parsing(check_status.id)
            log_event(e, filename, contractor, i)
            # print(e)
            return ({contractor: 'error', 'id': check_status.id})
    else:
        return ({contractor: 'Data is parsed today'})
