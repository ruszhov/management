import os
from django.conf import settings
import datetime
import openpyxl
import re
import logging
import subprocess
import shlex
from django.http import StreamingHttpResponse

import sys
import os.path
sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from parsing.models import *
from parsing.parsers.functions import *

def brok(filename, contractor):
    # exception = kw.pop('exception', Exception)
    check_status = File.objects.get(file__icontains=filename, uploaded_at__date=datetime.date.today())
    operator_id = operator_create('Брок', contractor)
    header_row = 7
    table_first_row = 8
    light_presence = 'є'
    sold_presence = ['занято', 'зайнято']
    rezerv_presence = 'резерв'

    if check_status.status == 1:
        try:
            if filename.endswith('.xls'):
                if filename.endswith('.xls'):
                    dir = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()))
                    output = subprocess.run(['parsing/parsers/convert.sh', dir, filename])
                    filepath = os.path.join(dir, filename + 'x')
                else:
                    filepath = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()), filename)
                wb = openpyxl.load_workbook(filepath)
                sheet = wb.active
                # get max row count
                max_row = sheet.max_row
                # get max column count
                max_column = sheet.max_column
                # iterate over all cells
                # iterate over all rows
                year_id = year_create()
                m = months_create()

                header_dictionary = {}
                # get header row by number of string
                headers = sheet[header_row]
                for i in headers:
                    header_dictionary[i.col_idx] = i.value

                # iterate trough headers keys and find col indexes

                kwargs = {
                    'kod_col': 'Код',
                    'city_col': 'Нас.пункт',
                    'address_col': 'Адреса',
                    'type_col': 'Тип',
                    'size_col': 'Формат',
                    'side_col': 'Ст.',
                    'light_col': 'Осв.',
                    'url_col': 'Фото',
                    'price_col': 'вартість без пдв'
                }
                args = ('січень', 'лютий', 'березень', 'квітень', 'травень', 'червень', 'липень', 'серпень', 'вересень', 'жовтень', 'листопад', 'грудень')
                col_idx = get_value_from_header_dict(header_dictionary, **kwargs)
                col_idx_by_month = get_month_value_from_header_dict(header_dictionary, *args)

                # iterating trough rows
                for i in range(table_first_row, max_row + 1):
                    city_tmp = None
                    media_type_tmp = None
                    size_tmp = None
                    side_tmp = None

                    kod = sheet.cell(row=i, column=col_idx['kod_col'][0]).value
                    if kod is not None and 'сітілайт' not in kod:

                        city = sheet.cell(row=i, column=col_idx['city_col'][0]).value
                        city_id = city_create(city) if city != '' and city != city_tmp and city is not None else None
                        city_tmp = city

                        raw_address = sheet.cell(row=i, column=col_idx['address_col'][0]).value
                        address = raw_address if raw_address != '' and raw_address is not None else False

                        media_type = sheet.cell(row=i, column=col_idx['type_col'][0]).value
                        media_type_id = media_type_create(
                            media_type) if media_type != '' and media_type != media_type_tmp and media_type is not None else None
                        media_type_tmp = media_type

                        size = sheet.cell(row=i, column=col_idx['size_col'][0]).value
                        if size != '' and size != size_tmp and size is not None:
                            s = size.replace(',', '.').split(' ')[0]
                            size_id = board_size_create(s)
                            size_tmp = size
                        else:
                            size_id = None

                        side = sheet.cell(row=i, column=col_idx['side_col'][0]).value
                        board_side_id = board_side_create(
                            side) if side != '' and side != side_tmp and side is not None else None
                        side_tmp = side

                        raw_light = sheet.cell(row=i, column=col_idx['light_col'][0]).value
                        light = True if raw_light == light_presence else False

                        hyperlink = sheet.cell(row=i, column=col_idx['url_col'][0]).hyperlink
                        url = hyperlink.target if hyperlink is not None else None

                        price = sheet.cell(row=i, column=col_idx['price_col'][0]).value
                        price_val = float(price) if price != '' and price is not None else None

                        # save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors, operator_id)
                        board_id = save_board(kod, url, None, city_id, None, address, media_type_id, size_id, board_side_id, light, None, None, None, operator_id)

                        for k, v in col_idx_by_month.items():
                            status = sheet.cell(row=i, column=col_idx_by_month[k][0]).value
                            status_id = save_status(status, sold_presence, rezerv_presence)
                            save_status_price(board_id, status_id, price_val, k)
                    else:
                        pass

                    # response = StreamingHttpResponse(i , content_type='application/json')
                    # return response
            done_parsing(check_status.id)
            return ({contractor: 'parsed'})
            #response = StreamingHttpResponse(i , content_type='application/json')
            #return response

        except Exception as e:
            error_parsing(check_status.id)
            log_event(e, filename, contractor, i)
            # print(e)
            return ({contractor: 'error', 'id': check_status.id})
    else:
        return ({contractor: 'Data is parsed today'})
