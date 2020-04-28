import os
from django.conf import settings
import datetime
import openpyxl
import re
import logging

import sys
import os.path
sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from parsing.models import *
from parsing.parsers.functions import *

def media_mist(filename, contractor):
    # exception = kw.pop('exception', Exception)
    check_status = File.objects.get(file__icontains=filename, uploaded_at__date=datetime.date.today())
    operator_id = operator_create('Медіа Міст', contractor)
    header_row = 4
    table_first_row = 7
    light_presence = ''
    sold_presence = 'З'
    rezerv_presence = ['Р', 'р']

    if check_status.status == 1:
        try:
            if filename.endswith('.xls'):
                dir = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()))
                output = subprocess.run(['parsing/parsers/convert.sh', dir, filename])
                filepath = os.path.join(dir, filename + 'x')
            else:
                filepath = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()), filename)

            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            # get max row count
            max_row = sheet.max_row
            # get max column count
            max_column = sheet.max_column
            # iterate over all cells
            # iterate over all rows
            year_id = year_create()
            m = months_create()
            header_dictionary = {}

            headers = sheet[header_row]
            for i in headers:
                if i.value is not None:
                    header_dictionary[i.col_idx] = i.value

            kwargs = {
                'kod_col' : 'Назва',
                'url_col' : 'URL',
                'region_col' : 'Область',
                'city_col' : 'Місто',
                'city_area_col' : 'Район міста',
                'address_col' : 'Розташування',
                'address_col_1' : 'Привязка до обектів інфраструктури',
                'type_col' : 'Тип',
                'size_col' : 'Розмір',
                'side_col' : 'Сторона',
                'light_col' : 'Підсвітка щита',
                'price_col' : 'Ціна оренди'
            }
            args = ('січ.', 'лют.', 'бер.', 'квіт.', 'трав.', 'чер.', 'лип.', 'серп.', 'вер.', 'жов.', 'лист.', 'груд.')

            col_idx = get_value_from_header_dict(header_dictionary, **kwargs)
            col_idx_by_month = get_month_value_from_header_dict(header_dictionary, *args)

            for i in range(table_first_row, max_row + 1):
                current_row = i
                region_tmp = None
                city_tmp = None
                city_area_tmp = None
                media_type_tmp = None
                size_tmp = None
                side_tmp = None
                regex = re.compile('[-@_!#$%^&*()<>=?/\|}{~:]')

                price = sheet.cell(row=i, column=col_idx['price_col'][0]).value
                raw_kod = sheet.cell(row=i, column=col_idx['kod_col'][0]).value
                if raw_kod is not None and raw_kod != '' and price is not None and price != '':

                    kod = raw_kod

                    hyperlink = sheet.cell(row=i, column=col_idx['url_col'][0]).value
                    if hyperlink is not None and hyperlink != '':
                        hyp_splited = str(hyperlink).split('"')
                        if len(hyp_splited) > 1:
                            url = hyp_splited[1]
                        else:
                            url = None

                    region = sheet.cell(row=i, column=col_idx['region_col'][0]).value
                    region_id = region_create(region) if region != '' and region != region_tmp and region is not None else None
                    region_tmp = region

                    city = sheet.cell(row=i, column=col_idx['city_col'][0]).value
                    city_id = city_create(city) if city != '' and city != city_tmp and city is not None else None
                    city_tmp = city

                    city_area = sheet.cell(row=i, column=col_idx['city_area_col'][0]).value
                    city_area_id = city_area_create(city_area) if city_area != '' and city_area != city_area_tmp and city_area is not None else None
                    city_area_tmp = city_area

                    address_1 = sheet.cell(row=i, column=col_idx['address_col'][0]).value
                    address_2 = sheet.cell(row=i, column=col_idx['address_col_1'][0]).value
                    if address_1 is not None and '=' not in address_1:
                        address = address_1 + ' ('+address_2+')'
                    elif address_2 is not None:
                        address = address_2
                    else:
                        address = None

                    side = sheet.cell(row=i, column=col_idx['side_col'][0]).value
                    board_side_id = board_side_create(side) if side != '' and side != side_tmp and side is not None else None
                    side_tmp = side

                    size = sheet.cell(row=i, column=col_idx['size_col'][0]).value
                    if size is not None:
                        if size != '' and regex.search(size) is None:
                            s = size.replace(',', '.').split(' ')[0]
                            size_id = board_size_create(s)
                    else:
                        size_id = None

                    media_type = sheet.cell(row=i, column=col_idx['type_col'][0]).value
                    media_type_id = media_type_create(media_type) if media_type != '' and media_type != media_type_tmp and media_type is not None else None
                    media_type_tmp = media_type

                    raw_light = sheet.cell(row=i, column=col_idx['light_col'][0]).value
                    light = True if raw_light == light_presence else False

                    price = sheet.cell(row=i, column=col_idx['price_col'][0]).value
                    price_val = float(price) if price != '' and price is not None else None

                    #### save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors, operator_id)
                    board_id = save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, None, None, None, operator_id)

                    for k, v in col_idx_by_month.items():
                        status = sheet.cell(row=i, column=col_idx_by_month[k][0]).value
                        status_id = save_status(status, sold_presence, rezerv_presence)
                        save_status_price(board_id, status_id, price_val, k)

                yield done_percent(i, max_row)

            # done_parsing(check_status.id)

        except Exception as e:
            error_parsing(check_status.id)
            log_event(e, filename, contractor, current_row)
            yield ({contractor: 'error', 'id': check_status.id})
    else:
        pass
