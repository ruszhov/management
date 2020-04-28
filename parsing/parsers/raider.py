import os
from django.conf import settings
import datetime
import openpyxl
import re
import logging
import subprocess

import sys
import os.path
sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from parsing.models import *
from parsing.parsers.functions import *

def raider(filename, contractor):
    # exception = kw.pop('exception', Exception)
    check_status = File.objects.get(file__icontains=filename, uploaded_at__date=datetime.date.today())
    operator_id = operator_create('Рейдер', contractor)
    if ('CityLight').lower() in filename.lower():
        header_row = 8
        table_first_row = 9
    else:
        header_row = 7
        table_first_row = 8

    light_presence = 'так'
    sold_presence = 'продано'
    rezerv_presence = 'резерв'

    if check_status.status == 1:
        try:
            if filename.endswith('.xls'):
                dir = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()))
                output = subprocess.run(['parsing/parsers/convert.sh', dir, filename])
                filepath = os.path.join(dir, filename + 'x')
            else:
                filepath = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()), filename)

            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            # get max row count
            max_row = sheet.max_row
            # get max column count
            max_column = sheet.max_column
            # iterate over all cells
            # iterate over all rows
            # year_id = year_create()
            m = months_create()
            free_status_id = free_status_create()
            header_dictionary = {}
            months_data = {}

            def add_to_dict(k, v):
                try:
                    months_data[k].append(v)
                except KeyError:
                    months_data[k] = [v]

            headers = sheet[header_row]
            for i in headers:
                if i.value is not None:
                    header_dictionary[i.col_idx] = i.value
            # iterate trough headers keys and find col indexes
            for k, v in header_dictionary.items():
                if not isinstance(v, datetime.datetime):
                    if 'Код' == v:
                        kod_col = k
                    if '№' == v:
                        num_col = k
                    elif any(c in v for c in ('Ст.', 'Сторона')):
                        side_col = k
                    elif 'Місто' in v:
                        city_col = k
                    elif 'Адреса' in v:
                        address_col = k
                    elif 'Прив\'язка до місця' in v:
                        address_1_col = k
                    elif 'Формат' in v:
                        size_col = k
                    elif any(c in v for c in ('Під-світка', 'Підсвітка', 'Підс\nвітка')):
                        light_col = k
                    elif 'Ціна' in v:
                        price_col = k
                else:
                    today = datetime.datetime.today()
                    if today.year == v.year:
                        add_to_dict(v.month, k)

            # iterating trough rows
            for i in range(table_first_row, max_row + 1):
                city_tmp = None
                size_tmp = None
                side_tmp = None

                raw_kod = sheet.cell(row=i, column=kod_col).value
                kod = raw_kod if raw_kod != '' and raw_kod is not None else None

                hyperlink = sheet.cell(row=i, column=kod_col).hyperlink
                url = hyperlink.target if hyperlink != '' and hyperlink is not None else None

                if kod !=None and kod != '' and len(kod) <= 10 and kod !='Сітілайти' and kod !='Скролери':

                    try:
                        kod += ' '+sheet.cell(row=i, column=num_col).value
                    except NameError:
                        pass

                    side = sheet.cell(row=i, column=side_col).value
                    board_side_id = board_side_create(side) if side != '' and side != side_tmp and side is not None else None
                    side_tmp = side

                    city = sheet.cell(row=i, column=city_col).value
                    city_id = city_create(city) if city != '' and city != city_tmp and city is not None else None
                    city_tmp = city

                    # address = sheet.cell(row=i, column=address_col).value +' '+sheet.cell(row=1, column=address_1_col).value
                    # print(address)
                    raw_address = sheet.cell(row=i, column=address_col).value
                    raw_address_1 = sheet.cell(row=i, column=address_1_col).value
                    address_1 = raw_address_1 if raw_address_1 != '' and raw_address_1 is not None else None
                    if address_1 is not None:
                        address = raw_address + ' (' + address_1 + ')' if raw_address != '' and raw_address is not None else None
                    else:
                        address = raw_address if raw_address != '' and raw_address is not None else None


                    size = sheet.cell(row=i, column=size_col).value
                    if size != '' and size != size_tmp and size is not None:
                        s = size.replace('*', 'x').replace(',', '.').split(' ')[0]
                        size_id = board_size_create(s)
                        size_tmp = size
                    else:
                        size_id = None

                    raw_light = sheet.cell(row=i, column=light_col).value
                    light = True if raw_light == light_presence else False

                    price = sheet.cell(row=i, column=price_col).value
                    price_val = float(price) if price != '' and price is not None else None

                    city_area_id = None
                    media_type_id = None
                    ots = None
                    grp = None
                    kod_doors=None
                    region_id=None

                    #board saving to DB
                    board_id = save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors, operator_id)

                    for k, v in months_data.items():
                        status = sheet.cell(row=i, column=v[0]).value
                        status_id = save_status(status, sold_presence, rezerv_presence)
                        save_status_price(board_id, status_id, price_val, k)

            done_parsing(check_status.id)
            return ({contractor: 'parsed'})

        except Exception as e:
            error_parsing(check_status.id)
            log_event(e, filename, contractor, i)
            # print(e)
            return ({contractor: 'error', 'id': check_status.id})
    else:
        return ({contractor: 'Data is parsed today'})
