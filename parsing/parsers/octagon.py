import os
from django.conf import settings
import datetime
import openpyxl
import re
import logging

import sys
import os.path
sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from parsing.models import *
from parsing.parsers.functions import *

def octagon(filename, contractor):
    # exception = kw.pop('exception', Exception)
    check_status = File.objects.get(file__icontains=filename, uploaded_at__date=datetime.date.today())
    operator_id = operator_create('Октагон', contractor)
    header_row = 2
    table_first_row = 3
    light_presence = 'Есть'
    sold_presence = 'Занято'
    rezerv_presence = ['Резерв', 'Резерв Резерв']

    if check_status.status == 1:
        try:
            if filename.endswith('.xls'):
                dir = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()))
                output = subprocess.run(['parsing/parsers/convert.sh', dir, filename])
                filepath = os.path.join(dir, filename + 'x')
            else:
                filepath = os.path.join(settings.MEDIA_ROOT, 'workflow', str(datetime.date.today()), filename)

            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            # get max row count
            max_row = sheet.max_row
            # get max column count
            max_column = sheet.max_column
            # iterate over all cells
            # iterate over all rows
            year_id = year_create()
            m = months_create()
            header_dictionary = {}

            headers = sheet[header_row]
            for i in headers:
                header_dictionary[i.col_idx] = i.value

            kwargs = {
                'kod_col': 'Код',
                'url_col': 'фото\\схема',
                'region_col': 'Область',
                'city_col': ' Город',
                'city_area_col': 'Район',
                'address_col': 'Адрес',
                'type_col': ' Формат',
                'side_col': 'Ст.',
                'light_col': 'Свет',
                'ots_col': 'OTS, тыс. чел.',
                'grp_col': 'GRP',
                'doors_col': 'Номер Doors',
                'price_col': 'Цена грн. без НДС'
            }
            args = ('Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек')
            col_idx = get_value_from_header_dict(header_dictionary, **kwargs)
            col_idx_by_month = get_month_value_from_header_dict(header_dictionary, *args)
            # iterating trough rows
            for i in range(table_first_row, max_row + 1):

                region_tmp = None
                city_tmp = None
                city_area_tmp = None
                media_type_tmp = None
                side_tmp = None
                address_tmp = None

                kod = sheet.cell(row=i, column=col_idx['kod_col'][0]).value
                if kod is not None and kod != '':

                    region = sheet.cell(row=i, column=col_idx['region_col'][0]).value
                    region_id = region_create(region) if region != '' and region != region_tmp and region is not None else None
                    region_tmp = region

                    city = sheet.cell(row=i, column=col_idx['city_col'][0]).value
                    city_id = city_create(city) if city != '' and city != city_tmp and city is not None else None
                    city_tmp = city

                    city_area = sheet.cell(row=i, column=col_idx['city_area_col'][0]).value
                    city_area_id = city_area_create(city_area) if city_area != '' and city_area != city_area_tmp and city_area is not None else None
                    city_area_tmp = city_area

                    media_type_raw = sheet.cell(row=i, column=col_idx['type_col'][0]).value
                    if media_type_raw != '' and media_type_raw != media_type_tmp and media_type_raw is not None:
                        str_split = media_type_raw.split()
                        media_type = ''
                        for l in str_split:
                            if l.isalpha():
                                media_type += l + ' '
                            else:
                                if '[' not in l:
                                    size = l
                                else:
                                    pass
                        media_type_tmp = media_type
                    else:
                        media_type = None
                        size = None

                    if size != '' and size is not None:
                        s = size.replace(',', '.').split(' ')[0]
                        size_id = board_size_create(s)
                    else:
                        size_id = None

                    media_type_id = media_type_create(media_type) if media_type != '' and media_type is not None else None

                    raw_light = sheet.cell(row=i, column=col_idx['light_col'][0]).value
                    light = True if raw_light == light_presence else False

                    side = sheet.cell(row=i, column=col_idx['side_col'][0]).value
                    board_side_id = board_side_create(side) if side != '' and side != side_tmp and side is not None else None
                    side_tmp = side

                    address = sheet.cell(row=i, column=col_idx['address_col'][0]).value
                    address = address if address != '' and address != address_tmp and address is not None else None
                    address_tmp = address
                    # kod = sheet.cell(row=i, column=kod_col).value
                    price = sheet.cell(row=i, column=col_idx['price_col'][0]).value
                    price_val = float(price) if price != '' and price is not None else None
                    hyperlink = sheet.cell(row=i, column=col_idx['url_col'][0]).hyperlink
                    url = hyperlink.target if hyperlink != '' and hyperlink is not None else None
                    ots = sheet.cell(row=i, column=col_idx['ots_col'][0]).value
                    ots = int(ots) if ots and ots !='' and ots is not None else None
                    raw_grp = sheet.cell(row=i, column=col_idx['grp_col'][0]).value
                    grp = float(raw_grp.replace(',', '.')) if raw_grp and raw_grp != '' and raw_grp is not None else None

                    kod_doors = sheet.cell(row=i, column=col_idx['doors_col'][0]).value
                    kod_doors_int = int(kod_doors) if kod_doors != '' and kod_doors is not None else None

                    # save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors, operator_id)
                    board_id = save_board(kod, url, region_id, city_id, city_area_id, address, media_type_id, size_id, board_side_id, light, ots, grp, kod_doors_int, operator_id)

                    for k, v in col_idx_by_month.items():
                        status = sheet.cell(row=i, column=col_idx_by_month[k][0]).value
                        status_id = save_status(status, sold_presence, rezerv_presence)
                        save_status_price(board_id, status_id, price_val, k)

            done_parsing(check_status.id)
            return ({contractor: 'parsed'})

        except Exception as e:
            error_parsing(check_status.id)
            log_event(e, filename, contractor, i)
            # print(e)
            return ({contractor: 'error', 'id': check_status.id})
    else:
        return ({contractor: 'Data is parsed today'})
